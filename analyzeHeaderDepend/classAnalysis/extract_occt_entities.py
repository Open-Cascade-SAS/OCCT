import sys
import os
import glob
import re
import clang.cindex
from clang.cindex import CursorKind
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# ================= 配置部分 =================
# 1. libclang 路径
CLANG_LIBRARY_FILE = r"E:\workFile\temp\analyzeHeaderDepend\classAnalysis\clang+llvm-21.1.8-x86_64-pc-windows-msvc\clang+llvm-21.1.8-x86_64-pc-windows-msvc\bin\libclang.dll"

# 2. OCCT 源码目录
OCCT_SRC_DIR = r"E:\code\occtSrc\OCCT\src"

# 3. OCCT 头文件目录
OCCT_INC_DIR = r"E:\code\occtSrc\Install\inc"

# 4. 输出文件
OUTPUT_EXCEL = "occt_entities_for_refactoring.xlsx"
# ===========================================

if not os.path.exists(CLANG_LIBRARY_FILE):
    print(f"Error: {CLANG_LIBRARY_FILE} not found.")
    sys.exit(1)
clang.cindex.Config.set_library_file(CLANG_LIBRARY_FILE)

def build_package_to_toolkit_map(src_dir):
    """
    解析所有 PACKAGES 文件，构建 Package -> Toolkit 映射
    """
    package_map = {}
    packages_files = glob.glob(os.path.join(src_dir, "*", "PACKAGES"))
    
    for pkg_file in packages_files:
        toolkit_name = os.path.basename(os.path.dirname(pkg_file))
        
        with open(pkg_file, 'r', encoding='utf-8') as f:
            for line in f:
                package_name = line.strip()
                if package_name:
                    package_map[package_name] = toolkit_name
    
    print(f"已解析 {len(packages_files)} 个 PACKAGES 文件，共 {len(package_map)} 个包")
    return package_map

def infer_package_from_filename(filepath):
    """
    从文件名推断 Package 名称
    """
    basename = os.path.basename(filepath)
    name_without_ext = os.path.splitext(basename)[0]
    
    if '_' in name_without_ext:
        package_name = name_without_ext.split('_')[0]
        return package_name
    else:
        return name_without_ext

def get_file_base_names(src_dir, inc_dir):
    """
    获取所有文件基础名（头文件和源文件同名视为一个）
    返回: dict { 'BaseName': {'has_hxx': bool, 'has_cxx': bool, 'package': str} }
    """
    files_info = defaultdict(lambda: {'has_hxx': False, 'has_cxx': False, 'package': ''})
    
    # 扫描头文件
    for hxx_file in glob.glob(os.path.join(inc_dir, "*.hxx")):
        basename = os.path.splitext(os.path.basename(hxx_file))[0]
        files_info[basename]['has_hxx'] = True
        files_info[basename]['package'] = infer_package_from_filename(hxx_file)
    
    # 扫描源文件
    for cxx_file in glob.glob(os.path.join(src_dir, "*", "*.cxx")):
        basename = os.path.splitext(os.path.basename(cxx_file))[0]
        files_info[basename]['has_cxx'] = True
        if not files_info[basename]['package']:
            files_info[basename]['package'] = infer_package_from_filename(cxx_file)
    
    print(f"共找到 {len(files_info)} 个唯一文件名")
    return files_info

def extract_entities_from_file(filepath):
    """
    从单个头文件中提取类、结构体、枚举、命名空间等实体
    """
    index = clang.cindex.Index.create()
    args = [
        '-x', 'c++', 
        f'-I{OCCT_INC_DIR}', 
        '-DWNT', 
        '-DStandard_EXPORT=', 
        '-DStandard_VkExport=',
        '-std=c++14'
    ]
    
    try:
        tu = index.parse(filepath, args=args, options=clang.cindex.TranslationUnit.PARSE_SKIP_FUNCTION_BODIES)
    except Exception as e:
        print(f"解析失败: {filepath}, 错误: {e}")
        return []

    entities = []
    target_filename = os.path.basename(filepath).lower()
    
    def visit(node, parent_namespace=""):
        # 只处理当前文件定义的实体
        if node.location.file:
            node_fname = os.path.basename(node.location.file.name).lower()
            if node_fname != target_filename:
                return
        
        # 类
        if node.kind == CursorKind.CLASS_DECL and node.is_definition():
            if node.spelling:
                full_name = f"{parent_namespace}::{node.spelling}" if parent_namespace else node.spelling
                entities.append({
                    'type': 'Class',
                    'name': node.spelling,
                    'full_name': full_name,
                    'namespace': parent_namespace
                })
        
        # 结构体
        elif node.kind == CursorKind.STRUCT_DECL and node.is_definition():
            if node.spelling:
                full_name = f"{parent_namespace}::{node.spelling}" if parent_namespace else node.spelling
                entities.append({
                    'type': 'Struct',
                    'name': node.spelling,
                    'full_name': full_name,
                    'namespace': parent_namespace
                })
        
        # 枚举
        elif node.kind == CursorKind.ENUM_DECL and node.is_definition():
            if node.spelling:
                full_name = f"{parent_namespace}::{node.spelling}" if parent_namespace else node.spelling
                entities.append({
                    'type': 'Enum',
                    'name': node.spelling,
                    'full_name': full_name,
                    'namespace': parent_namespace
                })
        
        # 命名空间
        elif node.kind == CursorKind.NAMESPACE:
            if node.spelling:
                new_namespace = f"{parent_namespace}::{node.spelling}" if parent_namespace else node.spelling
                entities.append({
                    'type': 'Namespace',
                    'name': node.spelling,
                    'full_name': new_namespace,
                    'namespace': parent_namespace
                })
                # 递归处理命名空间内的实体
                for child in node.get_children():
                    visit(child, new_namespace)
                return
        
        # 递归处理子节点
        for child in node.get_children():
            visit(child, parent_namespace)
    
    visit(tu.cursor)
    return entities

def create_excel_report(package_map, files_info, all_entities):
    """
    创建 Excel 报告
    """
    wb = openpyxl.Workbook()
    
    # 删除默认的 Sheet
    wb.remove(wb.active)
    
    # ========== Sheet 1: Toolkits ==========
    ws_toolkits = wb.create_sheet("Toolkits")
    ws_toolkits.append(["Original Name", "Refactored Name", "Type", "Package Count"])
    
    # 统计每个 Toolkit 的 Package 数量
    toolkit_packages = defaultdict(list)
    for package, toolkit in package_map.items():
        toolkit_packages[toolkit].append(package)
    
    for toolkit in sorted(toolkit_packages.keys()):
        ws_toolkits.append([
            toolkit,
            "",  # 重构后的名称（待填写）
            "Toolkit",
            len(toolkit_packages[toolkit])
        ])
    
    # ========== Sheet 2: Packages ==========
    ws_packages = wb.create_sheet("Packages")
    ws_packages.append(["Original Name", "Refactored Name", "Type", "Toolkit", "File Count"])
    
    # 统计每个 Package 的文件数量
    package_files = defaultdict(int)
    for basename, info in files_info.items():
        package = info['package']
        if package:
            package_files[package] += 1
    
    for package in sorted(package_map.keys()):
        ws_packages.append([
            package,
            "",  # 重构后的名称（待填写）
            "Package",
            package_map[package],
            package_files.get(package, 0)
        ])
    
    # ========== Sheet 3: Files ==========
    ws_files = wb.create_sheet("Files")
    ws_files.append(["Original Name", "Refactored Name", "Type", "Package", "Toolkit", "Has .hxx", "Has .cxx"])
    
    for basename in sorted(files_info.keys()):
        info = files_info[basename]
        package = info['package']
        toolkit = package_map.get(package, "Unknown")
        
        ws_files.append([
            basename,
            "",  # 重构后的名称（待填写）
            "File",
            package,
            toolkit,
            "Yes" if info['has_hxx'] else "No",
            "Yes" if info['has_cxx'] else "No"
        ])
    
    # ========== Sheet 4: Classes ==========
    ws_classes = wb.create_sheet("Classes")
    ws_classes.append(["Original Name", "Refactored Name", "Type", "Full Name", "Namespace", "Package", "Toolkit"])
    
    for entity in sorted(all_entities, key=lambda x: x['full_name']):
        if entity['type'] == 'Class':
            # 从类名推断 Package
            package = entity['name'].split('_')[0] if '_' in entity['name'] else entity['name']
            toolkit = package_map.get(package, "Unknown")
            
            ws_classes.append([
                entity['name'],
                "",  # 重构后的名称（待填写）
                entity['type'],
                entity['full_name'],
                entity['namespace'],
                package,
                toolkit
            ])
    
    # ========== Sheet 5: Structs ==========
    ws_structs = wb.create_sheet("Structs")
    ws_structs.append(["Original Name", "Refactored Name", "Type", "Full Name", "Namespace", "Package", "Toolkit"])
    
    for entity in sorted(all_entities, key=lambda x: x['full_name']):
        if entity['type'] == 'Struct':
            package = entity['name'].split('_')[0] if '_' in entity['name'] else entity['name']
            toolkit = package_map.get(package, "Unknown")
            
            ws_structs.append([
                entity['name'],
                "",
                entity['type'],
                entity['full_name'],
                entity['namespace'],
                package,
                toolkit
            ])
    
    # ========== Sheet 6: Enums ==========
    ws_enums = wb.create_sheet("Enums")
    ws_enums.append(["Original Name", "Refactored Name", "Type", "Full Name", "Namespace", "Package", "Toolkit"])
    
    for entity in sorted(all_entities, key=lambda x: x['full_name']):
        if entity['type'] == 'Enum':
            package = entity['name'].split('_')[0] if '_' in entity['name'] else entity['name']
            toolkit = package_map.get(package, "Unknown")
            
            ws_enums.append([
                entity['name'],
                "",
                entity['type'],
                entity['full_name'],
                entity['namespace'],
                package,
                toolkit
            ])
    
    # ========== Sheet 7: Namespaces ==========
    ws_namespaces = wb.create_sheet("Namespaces")
    ws_namespaces.append(["Original Name", "Refactored Name", "Type", "Full Name", "Parent Namespace"])
    
    for entity in sorted(all_entities, key=lambda x: x['full_name']):
        if entity['type'] == 'Namespace':
            ws_namespaces.append([
                entity['name'],
                "",
                entity['type'],
                entity['full_name'],
                entity['namespace']
            ])
    
    # 美化所有工作表
    for ws in wb.worksheets:
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
    
    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel 报告已保存至: {OUTPUT_EXCEL}")

def main():
    print("开始提取 OCCT 实体信息...")
    
    # 1. 构建 Package -> Toolkit 映射
    package_map = build_package_to_toolkit_map(OCCT_SRC_DIR)
    
    # 2. 获取所有文件信息
    files_info = get_file_base_names(OCCT_SRC_DIR, OCCT_INC_DIR)
    
    # 3. 从所有头文件中提取实体
    all_entities = []
    hxx_files = glob.glob(os.path.join(OCCT_INC_DIR, "*.hxx"))
    
    print(f"\n开始解析 {len(hxx_files)} 个头文件...")
    for i, hxx_file in enumerate(hxx_files, 1):
        if i % 500 == 0:
            print(f"进度: {i}/{len(hxx_files)} ({(i/len(hxx_files)*100):.1f}%)")
        
        entities = extract_entities_from_file(hxx_file)
        all_entities.extend(entities)
    
    print(f"\n共提取 {len(all_entities)} 个实体:")
    entity_counts = defaultdict(int)
    for entity in all_entities:
        entity_counts[entity['type']] += 1
    
    for entity_type, count in sorted(entity_counts.items()):
        print(f"  - {entity_type}: {count}")
    
    # 4. 创建 Excel 报告
    print("\n正在生成 Excel 报告...")
    create_excel_report(package_map, files_info, all_entities)
    
    print("\n完成！")

if __name__ == "__main__":
    main()
